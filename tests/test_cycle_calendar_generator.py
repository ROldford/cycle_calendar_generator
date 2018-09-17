#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""Tests for `cycle_calendar_generator` package."""


import unittest
from unittest import mock
import os
import datetime

from pyfakefs import fake_filesystem_unittest
import openpyxl

from cycle_calendar_generator import cycle_calendar_generator


class Test_get_args(unittest.TestCase):
  # Get directory from args
  """Tests function to get folder argument and give default if none given"""

  def setUp(self):
    """Set up test fixtures, if any."""

  def tearDown(self):
    """Tear down test fixtures, if any."""

  @mock.patch(
    'cycle_calendar_generator.cycle_calendar_generator.argparse.ArgumentParser.parse_args'
  )
  def test_if_arg_is_string(self, mock_parse_args):
    """Normal input on command line"""
    mock_parse_args.return_value = cycle_calendar_generator.argparse.Namespace(
        folder='string'
    )
    self.assertIsInstance(cycle_calendar_generator.getArgs(), str)

  @mock.patch(
    'cycle_calendar_generator.cycle_calendar_generator.argparse.ArgumentParser.parse_args'
  )
  @mock.patch('cycle_calendar_generator.cycle_calendar_generator.os.getcwd')
  def test_gets_current_dir_if_no_arg_given(self, mock_getcwd, mock_parse_args):
    """Use current working directory if no folder given"""
    cwd_string = 'current working directory'
    mock_parse_args.return_value = cycle_calendar_generator.argparse.Namespace(
        folder=None
    )
    mock_getcwd.return_value = cwd_string
    self.assertEqual(cycle_calendar_generator.getArgs(), cwd_string)

class Test_read_schedule_setup_file(fake_filesystem_unittest.TestCase):
  """Tests function to open and read schedule setup Excel file"""
  """Reading is assumed to work correctly"""
  """Tests only check that exceptions are raised properly on bad input"""
  test_folder_path = '/test-folder'
  setup_filepath = "{}/{}".format(
    test_folder_path, cycle_calendar_generator.SCHEDULE_SETUP_FILENAME
  )
  data_justATextDoc = "This isn't actually an Excel file."

  def setUp(self):
    self.setUpPyfakefs()
    os.mkdir(self.test_folder_path)

  def test_raises_valueerror_if_invalid_path(self):
    """If input string is not a valid folder path, raise ValueError"""
    self.assertRaisesRegex(
        ValueError,
        cycle_calendar_generator.ERROR_INVALID_FOLDER,
        cycle_calendar_generator.readScheduleSetupFile,
        '/test-notafolder'
    )

  def test_raises_valueerror_if_no_setup_file(self):
    """If no Excel file matching preset filename exists, raise ValueError"""
    self.assertRaisesRegex(
        ValueError,
        cycle_calendar_generator.ERROR_MISSING_SETUP_FILE,
        cycle_calendar_generator.readScheduleSetupFile,
        self.test_folder_path
    )

  def test_raises_valueerror_if_setup_file_not_excel(self):
    """If setup file isn't an Excel file, raise ValueError"""
    # just a text file
    self.assertTrue(os.path.exists(self.test_folder_path))
    with open(self.setup_filepath, "x") as file:
      file.write(self.data_justATextDoc)
    self.assertTrue(os.path.exists(self.setup_filepath))
    self.assertRaisesRegex(
      ValueError,
      cycle_calendar_generator.ERROR_SETUP_FILE_NOT_EXCEL,
      cycle_calendar_generator.readScheduleSetupFile,
      self.test_folder_path
    )

class Test_parse_schedule_setup_file(unittest.TestCase):
  # Parse schedule setup workbook input (checking for valid data)
  # Parsing should generate:
  ## [dict]periodTiming -> [int]periodNumber: [tuple(Date, Date)](startTime, endTime)
  ## [list]CycleDaysList -> (str)cycleDay {showing all cycleDay strings}
  ## [dict]yearlySchedule -> [Date]date: [str]cycleDay
  """Tests function to parse schedule setup Excel file"""
  # Setting up good and bad Excel files
  data_periodTiming = [
    ["Period Number", "Start Time", "End Time"],
    ["1", "08:00 AM", "09:00 AM"],
    ["2", "09:00 AM", "10:00 AM"],
    ["3", "10:00 AM", "11:00 AM"],
    ["4", "11:00 AM", "12:00 PM"],
    ["5", "12:00 PM", "01:00 PM"],
  ]
  parsed_periodTiming = {
    "1": ("08:00 AM", "09:00 AM"),
    "2": ("09:00 AM", "10:00 AM"),
    "3": ("10:00 AM", "11:00 AM"),
    "4": ("11:00 AM", "12:00 PM"),
    "5": ("12:00 PM", "01:00 PM"),
  }
  data_cycleDaysList = ["A1", "B2", "C3", "D4", "E5", "F6"]
  parsed_cycleDaysList = data_cycleDaysList
  data_yearlySchedule = [
    ["Date", "Cycle Day"],
    ["August 31", data_cycleDaysList[0]],
    ["September 3", data_cycleDaysList[1]],
    ["September 4", data_cycleDaysList[2]],
    ["September 5", data_cycleDaysList[3]],
    ["September 6", data_cycleDaysList[4]],
    ["September 7", data_cycleDaysList[5]],
  ]
  parsed_yearlySchedule = {
    "August 31": data_cycleDaysList[0],
    "September 3": data_cycleDaysList[1],
    "September 4": data_cycleDaysList[2],
    "September 5": data_cycleDaysList[3],
    "September 6": data_cycleDaysList[4],
    "September 7": data_cycleDaysList[5],
  }
  data_badExcel = [
    ["This", "isn't", "the", "right"],
    ["data", "for", "the", "parser"]
  ]
  wb_setup_good = openpyxl.Workbook()
  sheetname_periodTiming = "Period Timing"
  sheetname_cycleDaysList = "Cycle Days List"
  sheetname_yearlySchedule = "Yearly Schedule"
  ws_periodTiming = wb_setup_good.create_sheet(sheetname_periodTiming)
  ws_cycleDaysList = wb_setup_good.create_sheet(sheetname_cycleDaysList)
  ws_yearlySchedule = wb_setup_good.create_sheet(sheetname_yearlySchedule)
  for line in data_periodTiming:
    ws_periodTiming.append(line)
  ws_cycleDaysList.append(data_cycleDaysList)
  for line in data_yearlySchedule:
    ws_yearlySchedule.append(line)
  wb_setup_bad = openpyxl.Workbook()
  ws_bad = wb_setup_bad.active
  for line in data_badExcel:
    ws_bad.append(line)

  def test_parses_correct_setup(self):
    """Expected behavior: finds file with preset filename, opens and parses,
    returning object containing setup dicts and lists"""
    # self.assert? setup dicts and lists produced properly
    parsed_setup = cycle_calendar_generator.parseScheduleSetup(
      self.wb_setup_good
    )
    self.assertIsInstance(parsed_setup, cycle_calendar_generator.SetupData)
    self.assertEqual(self.parsed_periodTiming, parsed_setup.periodTiming)
    self.assertEqual(self.parsed_cycleDaysList, parsed_setup.cycleDaysList)
    self.assertEqual(self.parsed_yearlySchedule, parsed_setup.yearlySchedule)


  def test_raises_valueerror_if_setup_unparseable(self):
    """If Excel file can't be parsed following preset format, raise ValueError"""
    # test 1: just a text file
    self.assertRaisesRegex(
      ValueError,
      cycle_calendar_generator.ERROR_INVALID_SETUP_FILE,
      cycle_calendar_generator.parseScheduleSetup,
      self.wb_setup_bad
    )

class Test_teacher_schedule_file_scanner(unittest.TestCase):
  # Check for teacher schedule Excel files
  # Iterate over teacher schedule files
  """Tests function to scan through teacher schedule files and generate icals"""

class Test_read_teacher_schedule_file(fake_filesystem_unittest.TestCase):
  ## Read teacher schedule file
  ## Check if file is Excel file (exception check)
  """Tests function to read teacher schedule file and generate Workbook"""
  """Reading is assumed to work correctly"""
  """Tests only check that exceptions are raised properly on bad input"""
  test_folder_path = '/test-folder'
  teacher_filename = 'FirstnameLastname.xlsx'
  teacher_filepath = "{}/{}".format(
    test_folder_path, teacher_filename
  )
  teacher_filepath_bad = "{}/{}".format(
    'test-notafolder', teacher_filename
  )
  data_justATextDoc = "This isn't actually an Excel file."

  def setUp(self):
    self.setUpPyfakefs()
    os.mkdir(self.test_folder_path)

  def test_raises_valueerror_if_invalid_path(self):
    """If input string is not a valid folder path, raise ValueError"""
    self.assertRaisesRegex(
        ValueError,
        cycle_calendar_generator.ERROR_INVALID_SCHEDULE_FILE,
        cycle_calendar_generator.readTeacherScheduleFile,
        self.teacher_filepath_bad
    )

  def test_raises_valueerror_if_setup_file_not_excel(self):
    """If setup file isn't an Excel file, raise ValueError"""
    # just a text file
    self.assertTrue(os.path.exists(self.test_folder_path))
    with open(self.teacher_filepath, "x") as file:
      file.write(self.data_justATextDoc)
    self.assertTrue(os.path.exists(self.teacher_filepath))
    self.assertRaisesRegex(
      ValueError,
      cycle_calendar_generator.ERROR_INVALID_SCHEDULE_FILE,
      cycle_calendar_generator.readTeacherScheduleFile,
      self.teacher_filepath
    )

class Test_parse_teacher_schedule(unittest.TestCase):
  ## Check that file's periodNumbers (in first column) match those in setup file
  ## Check that file's cycleDays (in first row) match those in setup file
  ## Iterate over cycleDay columns, generating list of objects; for each...
  ### Generate dailySchedule object, 2 properties:
  #### [str]cycleDay
  #### [dict]schedule -> [int]periodNumber: [str]className
  ### Sort list by cycleDay property
  """Tests function to parse teacher schedule Workbook and make data object"""
  setupData_periodTiming =[
    ["1", "08:00 AM", "09:00 AM"],
    ["2", "09:00 AM", "10:00 AM"],
    ["3", "10:00 AM", "11:00 AM"],
    ["4", "11:00 AM", "12:00 PM"],
    ["5", "12:00 PM", "01:00 PM"],
  ]
  setupData_cycleDaysList = ["A1", "B2", "C3", "D4", "E5", "F6"]
  setupData_yearlySchedule = [
    ["Date", "Cycle Day"],
    ["August 31", setupData_cycleDaysList[0]],
    ["September 3", setupData_cycleDaysList[1]],
    ["September 4", setupData_cycleDaysList[2]],
    ["September 5", setupData_cycleDaysList[3]],
    ["September 6", setupData_cycleDaysList[4]],
    ["September 7", setupData_cycleDaysList[5]],
  ]
  wb_setup = openpyxl.Workbook()
  sheetname_periodTiming = "Period Timing"
  sheetname_cycleDaysList = "Cycle Days List"
  sheetname_yearlySchedule = "Yearly Schedule"
  ws_periodTiming = wb_setup.create_sheet(sheetname_periodTiming)
  ws_cycleDaysList = wb_setup.create_sheet(sheetname_cycleDaysList)
  ws_yearlySchedule = wb_setup.create_sheet(sheetname_yearlySchedule)
  for line in setupData_periodTiming:
    ws_periodTiming.append(line)
  ws_cycleDaysList.append(setupData_cycleDaysList)
  for line in setupData_yearlySchedule:
    ws_yearlySchedule.append(line)
  setupData = cycle_calendar_generator.parseScheduleSetup(
    wb_setup
  )
  data_teacherSchedule = [
    ["Period Number", "A1", "B2", "C3", "D4", "E5", "F6"],
    ["1", "Grade 8", "", "", "Grade 11", "", "Grade 8"],
    ["2", "", "Grade 8", "", "", "Grade 11", ""],
    ["3", "Lunch", "Lunch", "Lunch", "Lunch", "Lunch", "Lunch"],
    ["4", "Grade 11", "", "Grade 8", "", "", "Grade 11"],
    ["5", "", "Grade 11", "", "Grade 8", "", ""],
  ]
  parsed_teacherSchedule = cycle_calendar_generator.ScheduleData(
    ["1", "2", "3", "4", "5"]
  )
  parsed_teacherSchedule.addScheduleDay(
    ["A1", "Grade 8", "", "Lunch", "Grade 11", ""]
  )
  parsed_teacherSchedule.addScheduleDay(
    ["B2", "", "Grade 8", "Lunch", "", "Grade 11"]
  )
  parsed_teacherSchedule.addScheduleDay(
    ["C3", "", "", "Lunch", "Grade 8", ""]
  )
  parsed_teacherSchedule.addScheduleDay(
    ["D4", "Grade 11", "", "Lunch", "", "Grade 8"]
  )
  parsed_teacherSchedule.addScheduleDay(
    ["E5", "", "Grade 11", "Lunch", "", ""]
  )
  parsed_teacherSchedule.addScheduleDay(
    ["F6", "Grade 8", "", "Lunch", "Grade 11", ""]
  )
  data_badExcel = [
    ["This", "isn't", "the", "right"],
    ["data", "for", "the", "parser"]
  ]
  wb_schedule_good = openpyxl.Workbook()
  sheetname_teacherSchedule = "Teacher Schedule"
  ws_teacherSchedule = wb_schedule_good.create_sheet(sheetname_teacherSchedule)
  for line in data_teacherSchedule:
    ws_teacherSchedule.append(line)
  wb_schedule_bad = openpyxl.Workbook()
  ws_bad = wb_schedule_bad.active
  for line in data_badExcel:
    ws_bad.append(line)

  def test_parses_correct_schedule(self):
    """Expected behavior: finds file with preset filename, opens and parses,
    returning object containing schedule dicts / lists"""
    # self.assert? schedule dicts / lists produced properly
    parsed_schedule = cycle_calendar_generator.parseTeacherSchedule(
      self.wb_schedule_good, self.setupData
    )
    self.assertIsInstance(
      parsed_schedule, cycle_calendar_generator.ScheduleData
    )
    self.assertEqual(
      self.parsed_teacherSchedule, parsed_schedule.teacherSchedule
    )

  def test_raises_valueerror_if_schedule_unparseable(self):
    """If Excel file can't be parsed following preset format, raise ValueError"""
    self.assertRaisesRegex(
      ValueError,
      cycle_calendar_generator.ERROR_INVALID_SCHEDULE_FILE,
      cycle_calendar_generator.parseTeacherSchedule,
      self.wb_schedule_bad, self.setupData
    )

  def test_raises_valueerror_if_period_numbers_dont_match(self):
    schedule_sheet = self.wb_schedule_good["Teacher Schedule"]
    schedule_sheet["A6"] = "6"
    self.assertRaisesRegex(
      ValueError,
      cycle_calendar_generator.ERROR_INVALID_SCHEDULE_FILE,
      cycle_calendar_generator.parseTeacherSchedule,
      self.wb_schedule_good, self.setupData
    )

  def test_raises_valueerror_if_cycle_days_dont_match(self):
    schedule_sheet = self.wb_schedule_good["Teacher Schedule"]
    schedule_sheet["G1"] = "G7"
    self.assertRaisesRegex(
      ValueError,
      cycle_calendar_generator.ERROR_INVALID_SCHEDULE_FILE,
      cycle_calendar_generator.parseTeacherSchedule,
      self.wb_schedule_good, self.setupData
    )

class Test_generate_teacher_schedule_ical(unittest.TestCase):
  ## Create new iCal Calendar object
  ## Iterate over date:cycleDay dict, for each...
  ### Find dailySchedule object matching cycleDay
  ### Iterate over periodNumbers, for each...
  #### Check if className exists for this periodNumber, skip this if not
  #### Generate iCal Event object
  ##### Name of event = className
  ##### Start date, end date = this date
  ##### Start time, end time = found by referencing periodTiming
  #### Append Event object to Calendar
  """Tests function to use data object to make ical object"""

class Test_save_teacher_schedule_ical(unittest.TestCase):
  ## Save Calendar, using filename from teacher schedule file
  """Tests function to save teacher schedule ical file"""
